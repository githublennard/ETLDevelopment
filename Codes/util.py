# coding=utf8
import boto3
import time
import sys
import json
from pyspark import SparkContext, SQLContext, SparkConf
from pyspark.sql import Row, SparkSession, Column
from pyspark import sql
import numpy as np
import pyspark.sql.functions as F
from pyspark.sql.functions import isnan, when, count, col, current_timestamp, lit, monotonically_increasing_id, row_number, lower
import config.database_config as database_config
import config.config as config
import math
from database_utils import *
import traceback
from pyspark.sql.types import BooleanType, LongType, IntegerType, StringType, DateType, StructField, StructType, ArrayType, DoubleType, TimestampType
import io
import pandas as pd
from datetime import datetime, timedelta
import yaml
import base64
from botocore.exceptions import ClientError
import openpyxl



# Sin esto da error al hacer print con determinados caracteres
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

spark = SparkSession \
    .builder \
    .appName("DATAWALL: UTIL") \
    .getOrCreate()


def timestamp2seconds(ts):
    """ 
    Descripción: Convierte el valor tipo fecha en tipo numérico. Indica el número de segundos que han pasado desde el 1-1-1970 hasta la fecha correspondiente.
    
    Referencias:
        - Control_Outliers.py
        - Control_Rangos.py
    """
    years = ts.tm_year * 365 * 24 * 3600
    months = (ts.tm_mon - 1) * ((365./12.) * 24 * 3600)
    
    days = (ts.tm_mday - 1) * (24 * 3600) 
    
    hours = ts.tm_hour * 3600
    minutes = ts.tm_min * 60
    seconds = ts.tm_sec

    return years + months + days + hours + minutes + seconds 

def seconds2timestamp(secs ):    
    """ 
    Descripción: Convierte el valor tipo numérico en tipo fecha. Sabiendo que indica el número de segundos que han pasado desde el 1-1-1970 hasta la fecha correspondiente.
    
    Referencias:
        - Control_Ouliers.py
        - Control_Rangos.py
    """
    posix = secs - 62167219200 # Segundos de la fecha menos los segundos que hay entre el año 0 y el 1-1-1970 
    secs_in_year = 365 * 24 * 3600
    years_to_substract = 0

    while posix < 0: 
        posix += secs_in_year
        years_to_substract += 1

    ts = datetime.fromtimestamp(0)
    ts = ts.replace(year= ts.year - years_to_substract)

    return ts


def obtener_columna(df, col):    
    """ 
    Descripción: Dada un dataframe y una columna, devuelve los valores que tiene ese atributo en el dataframe. 
    
    Referencias:
    """
    return df.select(col).rdd.map(lambda r: r[0]).collect()

def query_FAST(query):
    """ 
    Descripción: Dada una query, devuelve el resultado obtenido realizando esta consulta en la BBDD de FAST de PRE.
    
    Referencias:
        - Control_Datatype.py
        - Control_Defaults.py
        - Control_Distinct.py
        - Control_DuplicadosInstanciasFAST.py
        - Control_DuplicadosServiciosFAST.py
        - Control_corruptos.py
        - Regla_VM_Fechas.py
        - Regla_VM_SLAs_Instancias.py
        - Regla_VM_SLAs_Ordenes.py
        - Warning_Correct
        - Control_CuentasHolding
    """
    if config.ENTORNO == "PRE":
        connection_data= json.loads(get_secret("pre-dq-rdb-fast"))
    else:
        connection_data= json.loads(get_secret("pro-dq-rdb-fast"))
    data_FAST = spark.read.format("jdbc") \
        .option("url", "jdbc:postgresql://" + connection_data["host"] + ":" + str(connection_data["port"]) + "/" + connection_data["dbname"]) \
        .option("user",connection_data["username"]) \
        .option("password", connection_data["password"])\
        .option("query", query) \
        .load()
    return data_FAST

def query_DQA(query):
    """ 
    Descripción: Dada una query, devuelve el resultado obtenido realizando esta consulta en la BBDD de DQA de PRE.
    
    Referencias:
        - Control_ConsistenciaFASTSF.py
        - Control_Datatype.py
        - Control_Defaults.py
        - Control_Distinct.py
        - Control_DuplicadosClienteFAST.py
        - Control_DuplicadosClienteSF.py
        - Control_KQIs.py
        - Control_MantenimientoOperacionesDataWall.py
        - Control_Transversal.py
        - Control_corruptos.py
        - Warning_Correct.py
        - Warning_SobrepasoUmbral.py
        - Warning_VariabilidadDiariaATR.py
        - warning_equipoPreventa.py
        - warning_mnc.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py
        - warning_wariabilidad_diaria.py
        - Control_ConsistenciaCuentasCoincidentesFastSF.py

    """
    if config.ENTORNO == "PRE":
        connection_data= json.loads(get_secret("pre-dq-rdb-dqa"))
    else:
        connection_data= json.loads(get_secret("pro-dq-rdb-dqa"))
    data_DQA = spark.read.format("jdbc") \
        .option("url", "jdbc:postgresql://" + connection_data["host"] + ":" + str(connection_data["port"]) + "/" + connection_data["dbname"]) \
        .option("user", connection_data["username"]) \
        .option("password", connection_data["password"])\
        .option("query", query) \
        .load()
    return data_DQA

def read_FAST(tabla):
    """ 
    Descripción: Devuelve un dataframe de la tabla que queramos de FAST que está en la BBDD de postgreSQL de PRE.
    
    Referencias:
        - Control_ConsistenciasFASTSF.py
        - Control_DuplicadosClienteFAST.py
        - Control_DuplicadosProveedoresFAST.py
        - Control_Transversal.py
        - Control_VendorFAST_V2.py
        - Regla_VM_Fechas.py
        - Regla_VM_SLAs_Instancias.py
        - Regla_VM_SLAs_Ordenes.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py
        - Control_ConsistenciaCuentasCoincidentesFastSF.py
    """

    data_FAST = query_FAST( "SELECT * FROM \""+tabla+"\"") 
    return data_FAST

def query_S3(path):
    """ 
    Descripción: Devuelve un dataframe del archivo que se encuentra en el path de S3 indicado.
    
    Referencias:
        - Control_ConsistenciaFASTSF.py
        - Control_CuentaOportunidad.py
        - Control_CuentaOportunidad_Holding.py
        - Control_DuplicadosClienteFAST.py
        - Control_DuplicadosClienteSF.py
        - Control_Tax_identification.py
        - Control_VendorFAST_V2.py
        - warning_equipoPreventa.py
    """

    data_S3 = spark.read.format("csv").option("sep", ";").option("header", "true").option(
        "inferSchema", "true").option("encoding", "UTF-8").load(path)
    return data_S3

def query_S32(path):

    data_S3 = spark.read.format("csv").option("sep", ",").option("header", "true").option(
        "inferSchema", "true").option("encoding", "UTF-8").load(path)
    return data_S3

def read_SF(tabla):
    """ 
    Descripción: Devuelve un dataframe de la tabla indicada de SalesForce que está en el bucket de S3.
    
    Referencias:
        - Control_ConsistenciaFASTSF.py
        - Control_CuentaOportunidad.py
        - Control_CuentaOportunidad_Holding.py
        - Control_Distinct.py
        - Control_DuplicadosClientesSF.py
        - Control_Transversal.py
        - Control_unicidad_contacto.py
        - Warning_equipoPreventa.py
        - Control_ConsistenciaCuentasCoincidentesFastSF.py
    """

    path = database_config.BUCKET_SF + tabla+".csv"
    data_SF = spark.read.format("csv").option("sep", ";").option("header", "true").option(
        "inferSchema", "true").option("encoding", "UTF-8").load(path)
    return data_SF

def sacarValores(vector):
    valores = []
    for i in range(len(vector)):
        valores.append(vector[i].asDict().values()[0])
    return valores

def getValues(dataframe):
    c = dataframe.collect()
    return [x[0] for x in c]

def read_RoD(tabla,fecha):
    """
    Descripción: Devuelve el fichero de la tabla de Remedy dada en la fecha indicada.

    Referencias:

    """
    # Leemos el dataframe del bucket de S3
    path=database_config.BUCKET_ROD +fecha+"/RD_TR_2_"+fecha+"_"+tabla+".txt"
    data_RoD = spark.read.format("csv").option("sep", "|" ).option("header", "false").option("inferSchema", "true").option("encoding","UTF-8").load(path)
    # Obtenemos los nombres de las columnas
    s3 = boto3.client('s3')
    obj = s3.get_object(Bucket='pro-dataquality-it',Key='00-Staging-Area/temp/Atributos_RoD.xlsx')
    #
    # Read using Pandas:
    # data_ColumnNames = obj['Body'].read()
    # df_ColumnNames = pd.read_excel(io.BytesIO(data_ColumnNames), sheet_name = tabla ,encoding='utf-8')
    # listOfLists = df_ColumnNames.values.tolist()
    # listColumnNames = [str(x[0]) for x in listOfLists]
    #
    # Read using Openpyxl
    data_excel = obj['Body'].read()
    wb = openpyxl.load_workbook(io.BytesIO(data_excel))
    ws = wb[tabla]
    listColumnNames = []
    for row in ws.iter_rows(min_row=ws.min_row + 1,max_row=ws.max_row):
        for cell in row:
            listColumnNames.append(cell.value)
    # Añadimos los nombre de las columnas al dataframe
    for i,c in enumerate(data_RoD.columns):
        data_RoD = data_RoD.withColumnRenamed(c,listColumnNames[i])
    return data_RoD

def read_json_SF_date(tabla,fecha):
    """ 
    Descripción: Devuelve la tabla de SalesForce dada en la fecha indicada. 
    
    Referencias:
    
    """
    dia = fecha.day
    mes = fecha.month
    anio = fecha.year
    s3client = boto3.client(
    's3',
    region_name='eu-central-1'
    )
    
    # These define the bucket and object to read
    bucketname = 'telefonicaiws-salesforce'
    file_to_read = "data/daily/" + str(anio) + "/" + str(mes) + "/" + str(dia) + "/" + tabla + ".json"

    #Create a file object using the bucket and object key. 
    fileobj = s3client.get_object(
        Bucket=bucketname,
        Key=file_to_read
        ) 
    # open the file object and read it into the variable filedata. 
    filedata = fileobj['Body'].read()

    # file data will be a binary stream.  We have to decode it 
    contents = filedata.encode("utf-8", errors="ignore").decode()
    data = checkDataYaml(contents)
    data = data["records"]

    df_final_spark = spark.sparkContext.parallelize(data).toDF(sampleRatio=1)

    return df_final_spark

def checkDataYaml(contents):
    # Elimina todos los caracteres no imprimibles
    while(True):
        try:
            data = yaml.safe_load(contents)
        except yaml.YAMLError as e:
            # print(e)
            contents = replace_str_index(contents, e.position)
            #data = yaml.safe_load(contents)
        except yaml.reader.ReaderError as ex:
            # print(e)
            contents = replace_str_index(contents, ex.position)
        else:
            return data
        

def replace_str_index(text,index=0,replacement=''):
    return '%s%s%s'%(text[:index],replacement,text[index+1:])

def read_json_SF(tabla):
    """ 
    Descripción: Devuelve la tabla de SalesForce dada en la fecha actual.
    
    Referencias:
    
    """
    fecha = datetime.now()
    dia = fecha.day-1
    mes = fecha.month
    anio = fecha.year
    s3client = boto3.client(
    's3',
    region_name='eu-central-1'
    )

    # These define the bucket and object to read
    # bucketname = 'telefonicaiws-salesforce'
    # file_to_read = "data/daily/" + str(anio) + "/" + str(mes) + "/" + str(dia) + "/" + tabla + ".json"
    bucketname = 'test-dqa'
    file_to_read = "pruebasJSON/" + tabla + ".json"

    #Create a file object using the bucket and object key. 
    fileobj = s3client.get_object(
        Bucket=bucketname,
        Key=file_to_read
        ) 
    # open the file object and read it into the variable filedata. 
    filedata = fileobj['Body'].read()

    # file data will be a binary stream.  We have to decode it 
    contents = filedata.encode("utf-8", errors="ignore").decode()

    data = checkDataYaml(contents)
    data = data["records"]

    df_final_spark = spark.sparkContext.parallelize(data).toDF(sampleRatio=1)

    return df_final_spark
    
def read_n_days_daily_SF(tabla, dias):
    """
        Descripcion:
            La funcion hace lo siguiente:
            1. Recoge los datos de los ultimos n dias 
            2. Les añade una columna del dia que se insertaron en SF de la tabla específica.
            3. Devuelve todos los datos recogidos en un DataFrame

        Parametros entrada:
            - tabla: Tabla sobre la que se va a recoger los datos de los ficheros JSON de SF
            - dias: Indica el número de días que se van a recoger
    """

    # date_select = datetime.strptime('2021-3-20', '%Y-%m-%d')
    # delta = timedelta(days=1)
    # fecha = date_select + delta

    fecha = datetime.now()
    jsonFile_df_final = None
    while dias - 1 > 0:
        dia = fecha-timedelta(days=dias-1)
        print(dia)
        jsonFile_df = read_json_SF_date(tabla,dia)
        jsonFile_df = jsonFile_df.withColumn("Fecha",lit(dia))
        #jsonFile_df.show()
        if jsonFile_df_final == None:
            jsonFile_df_final = jsonFile_df
        else:
            jsonFile_df_final, error, listColumns, listDeleteColumns = checkSchemas(jsonFile_df_final, jsonFile_df)
            if error == 1:
                print("En la tabla {0} del dia {1} se han creado las columnas {2}".format(tabla, str(dia.strftime('%Y-%m-%d')), listColumns))
            elif error == 2:
                print("En la tabla {0} del dia {1} se han eliminado las columnas {2}".format(tabla, str(dia.strftime('%Y-%m-%d')), listDeleteColumns))
            elif error == 3:
                print("En la tabla {0} del dia {1} se han creado las columnas {2} y se han eliminado las columnas {3}".format(tabla, str(dia.strftime('%Y-%m-%d')), listColumns, listDeleteColumns))
        dias = dias - 1
    return jsonFile_df_final

def checkSchemas(dfInicial, dfNuevo):

    elemPS = []
    elemNS = []
    error = 0
    listNewColumn = []
    listDeleteColumn  = []
    # Comprueba que las claves de los dos diccionarios estén incluidas en los dos diccionarios.
    schemaPS = dfInicial.schema.jsonValue()
    schemaNS = dfNuevo.schema.jsonValue()
    valuesPS = schemaPS.get("fields")
    valuesNS = schemaNS.get("fields")

    # Values es una lista que contiene solo los nombres de las columnas de la tabla
    for dic in valuesNS:
        elemNS.append(dic.get("name"))

    for dic in valuesPS:
        elemPS.append(dic.get("name"))



    for i, elem in enumerate(elemPS):
        if elem not in elemNS:
            dfNuevo = dfNuevo.withColumn(elem, lit(None))
            error = 2
            listDeleteColumn.append(elem)
            elemNS.append(elem)
            valuesNS.append(valuesNS[i])
            #valuesNS[clave] = valuesPS.get(clave)


    for i, elem in enumerate(elemNS):
        if elem not in elemPS:
            if error == 0:
                error = 1
            else:
                error = 3
            listNewColumn.append(elem)
            dfInicial = castColumn(dfInicial, elem, str(valuesNS[i].get("type")))
            elemPS.append(elem)
            valuesPS.append(valuesPS[i])


    dfInicial = dfInicial.select(sorted(dfInicial.columns))
    dfNuevo = dfNuevo.select(sorted(dfNuevo.columns))
    dfFinal = dfInicial.union(dfNuevo)

    return dfFinal, error, listNewColumn, listDeleteColumn

def castColumn (df, column, typeColumn):
    if typeColumn == "boolean":
        return df.withColumn(column, lit(None).cast(BooleanType()))
    elif typeColumn == "string":
        return df.withColumn(column, lit(None).cast(StringType()))
    elif typeColumn == "double":
        return df.withColumn(column, lit(None).cast(DoubleType()))
    elif typeColumn == "timestamp":
        return df.withColumn(column, lit(None).cast(TimestampType()))
    elif typeColumn == "integer":
        return df.withColumn(column, lit(None).cast(IntegerType()))
    elif typeColumn == "long":
        return df.withColumn(column, lit(None).cast(LongType()))
    elif typeColumn == "date":
        return df.withColumn(column, lit(None).cast(DateType()))
    else:
        return df.withColumn(column, lit(None))

def read_SF_parquet(tabla):
    sc = SQLContext(spark)
    path = database_config.RUTA_SALESFORCE + tabla + ".parquet"
    df = sc.read.parquet(path)
    return df


def df_2_db(df, tabla):
    if config.ENTORNO == "PRE":
        connection_data= json.loads(get_secret("pre-dq-rdb-dqa"))
    else:
        connection_data= json.loads(get_secret("pro-dq-rdb-dqa"))
    
    df.write \
        .format("jdbc") \
        .option("url", "jdbc:postgresql://" + connection_data["host"] + ":" + str(connection_data["port"]) + "/" + connection_data["dbname"]) \
        .option("user",connection_data["username"]) \
        .option("password", connection_data["password"])\
        .option("dbtable", tabla) \
        .mode("append")\
        .save()

def write_df_to_csv(dataframe, pathfile): 
    """
    Descripción: Almacena el dataframe con formato CSV en el path correspondiente.

    Referencias:
        - Warning_Correct.py
        - Warning_SobrepasoUmbral.py
        - Warning_VariabilidadDiariaATR.py
        - warning_mnc.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py

    """   
    dataframe.repartition(1).write.save(path=pathfile, format='csv', header = 'TRUE', mode='append', sep=';', emptyValue='NULL', float_format='{:19f}'.format, encoding='utf-8')

def calculate_Numreg(func):
    def inner(*args, **kwargs):
        if len(args) == 3 or kwargs.get("NumRegistros") is None:
            kwargs["NumRegistros"] = args[0].count()
        func(*args, **kwargs)
    return inner

@calculate_Numreg

def write_one_to_many_in_db(df_final, tabla, ClTipControl, ClAtrTabSis=None, NumRegistros=None, ListaClAtrTabSis=None):
    """ 
    Descripción: Escribe el DataFrame en la tabla de la BBDD de DQA. 

    Parámetros:
        · df_final -> DataFrame que se quiere almacenar.
        · tabla -> Nombre de la Tabla de DQA donde se quiere almacenar.
        · ClTipControl -> La clave correspondiente al control en la tabla "TipoControles" de DQA.
        · ClAtrTabSis -> Si el control está asociado a un único atributo, la clave de este.
        · NumRegistros -> El número de registros del DataFrame.
        · ListaClAtrTabSis -> Si el control está asociado a varios atributos, la clave de estos.

    Referencias:
        - Control_ConsistenciaFASTSF.py
        - Control_CuentaOportunidad.py
        - Control_CuentaOportunidad_Holding.py
        - Control_DuplicadosClientesFAST.py
        - Control_DuplicadosClientesSF.py
        - Control_KQIs.py
        - Control_MantenimientoOperacionDatawall.py
        - Control_Tax_identification.py
        - Control_unicidad_contacto.py
        - Regla_VM_Fechas.py
        - Regla_VM_SLAs_Instancias.py
        - Regla_VM_SLAs_Ordenes.py
        - Warning_Correct.py
        - Warning_SobrepasoUmbral.py
        - Warning_VariabilidadDiariaATR.py
        - warning_equipoPreventa.py
        - warning_mnc.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py
        - warning_wariabilidad_diaria.py
        - Control_ConsistenciaCuentasCoincidentesFastSF.py
        - Control_CuentasHolding
    """
    
    tabla_ejecucion_control = '"DQA"."EjecucionControl"'
    # Calcula la siguiente clave a partir de las que ya hay en la tabla "Ejecución Control" de DQA
    ClEjecControl = siguiente_clave(tabla_ejecucion_control)

    # Se crea el esquema de ejecución del control 
    df_EjecControl_schema = StructType([
        StructField("ClEjecControl", LongType(), False),
        StructField("ClAtrTabSis", IntegerType(), True),
        StructField("ClTipControl", IntegerType(), False),
        StructField("NumRegistros", LongType()),
        StructField("ListaClAtrTabSis", ArrayType(IntegerType()))
    ])

    # Se crea la tabla de ejecución del control 
    df_EjecucionControl = spark.createDataFrame(
        [[ClEjecControl, ClAtrTabSis, ClTipControl, NumRegistros,ListaClAtrTabSis]],
        df_EjecControl_schema)

    df_write_ejecucion = df_EjecucionControl.withColumn(
        "Fecha", current_timestamp())

    # Escribir el dataframe de ejecución del control en la tabla "EjecucionControl" de DQA
    df_2_db(df_write_ejecucion, '"DQA"."EjecucionControl"')

    # Escribir el dataframe en la tabla correspondiente de DQA
    df_write_control = df_final.withColumn(
        "ClEjecControl", lit(ClEjecControl))
    
    if "ClAtrTabSis" in df_write_control.columns:
        df_write_control = df_write_control.drop("ClAtrTabSis")
    if "NumRegistros" in df_write_control.columns:
        df_write_control = df_write_control.drop("NumRegistros")
    if "Fecha" in df_write_control.columns:
        df_write_control = df_write_control.drop("Fecha")
    if "ClTipControl" in df_write_control.columns:
        df_write_control = df_write_control.drop("ClTipControl")

    df_2_db(df_write_control, tabla)

def write_one_to_one_in_db(df_final, tabla, ClTipControl):
    """ 
    Descripción: Escribe el DataFrame en la tabla de la BBDD de DQA registro a registro. 

    Parámetros:
        · df_final -> DataFrame que se quiere almacenar.
        · tabla -> Nombre de la Tabla de DQA donde se quiere almacenar.
        · ClTipControl -> La clave correspondiente al control en la tabla "TipoControles" de DQA.

    Referencias:
        - Control_KQIs.py
        - Control_Nulos.py
        - Control_Rangos.py
        - Control_Transversal.py
        - warning_mnc.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py
    """
    tabla_ejecucion_control = '"DQA"."EjecucionControl"'

    # Obtenemos el número de filas en la tabla para obtener la siguiente clave ya que es incremental, asumimos que no van a existir condiciones de carrera
    start = siguiente_clave(tabla_ejecucion_control)
    df_write = crear_claves(df_final, "ClEjecControl", start)

    # Creación del dataframe de escritura de ejecución control
    if "NumRegistros" not in df_write.columns:
        func_num_registros = F.udf(lambda x: None,IntegerType())
        df_write = df_write.withColumn("NumRegistros",func_num_registros("ClEjecControl")) 
    df_write = df_write.withColumn(
        "Fecha", current_timestamp()).withColumn("ClTipControl", lit(ClTipControl))

    # Escritura del dataframe de ejecución control en la tabla "EjecucionControl" de DQA
    df_write_ejecucion = df_write["ClEjecControl",
                                  "ClAtrTabSis", "ClTipControl", "NumRegistros", "Fecha"]
    df_2_db(df_write_ejecucion, '"DQA"."EjecucionControl"')

    # Escritura del df_write en la tabla correspondiente
    df_write_control = df_write.drop("ClAtrTabSis","NumRegistros","Fecha","ClTipControl")
    df_2_db(df_write_control, tabla)

def pasar_excepciones(tipoControl, df_salida, tipoExcepcion):
    """
    Descripción: Elimina las excepciones indicadas en la tabla "Filtro" de DQA.

    Parámetros:
        · tipoControl -> La clave del control correspondiente a la tabla "TipoControles" de DQA.
        · df_salida -> DataFrame al que se le pasan las excepciones.
        · tipoExcepciones -> Clave del tipo de excepción.

    Referencias:
        - Control_ConsistenciaFASTSF.py
        - Control_CuentaOportunidad.py
        - Control_CuentaOportunidad_Holding.py
        - Control_DuplicadosClientesFAST.py
        - Control_DuplicadosClientesSF.py
        - Control_DuplicadosProveedoresFAST.py
        - Control_DuplicadosServiciosFAST.py
        - Control_DuplicadosServiciosSF.py
        - Control_Tax_identification.py
        - Control_unicidad_contacto.py
        - Regla_VM_Fechas.py
        - Regla_VM_SLAs_Instancias.py
        - Regla_VM_SLAs_Ordenes.py
        - Warning_Correct.py
        - Warning_VariabilidadDiariaATR.py
        - warning_equipoPreventa.py
        - warning_mnc.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py
    """
    # Se obtienen las excepciones correspondientes al control y al tipo de excepción
    query_excep = """
            SELECT "Atributo1", "Valor1", "Atributo2", "Valor2"
            FROM "DQA"."Filtro" fil 
                JOIN "DQA"."Excepcion" exc 
                    ON fil."ClExcepcion" = exc."ClExcepcion"
                JOIN "DQA"."RelExcepcionControl" rel
                    ON rel."ClExcepcion" = exc."ClExcepcion"
                JOIN "DQA"."TiposControles" tip
                    ON tip."ClTipControl" = rel."ClTipControl"
                JOIN "DQA"."TiposExcepciones" tipex
                    ON tipex."ClTipoExcepcion" = exc."ClTipoExcepcion"
            WHERE tip."ClTipControl" = """ + str(tipoControl) + """
                AND tipex."ClTipoExcepcion" = """ + str(tipoExcepcion) + """
                """
    df_excep = query_DQA(query_excep)
    df_return = df_salida
    if df_excep.count > 0:
        list_excep = df_excep.collect()
        list_valor1 = []
        list_valor2 = []
        list_attr1 = []
        list_attr2 = []
        for i in range(len(list_excep)):
            list_valor1.append(list_excep[i]['Valor1'])
            list_valor2.append(list_excep[i]['Valor2'])
            list_attr1.append(list_excep[i]['Atributo1'])
            list_attr2.append(list_excep[i]['Atributo2'])
        for j in range(len(list_attr1)):
            if (list_attr2[j] == None):
                # Filtrar el dataframe para que no salgan las excepciones
                df_return = df_return.filter(~F.col(list_attr1[j]).isin(list_valor1[j]))
            else:
                # Filtrar el dataframe para que no salgan las excepciones
                df_return = df_return.filter(~( (F.col(list_attr1[j]).isin(list_valor1[j])) & (F.col(list_attr2[j]).isin(list_valor2[j])) ))
    return df_return

def anadirExcepciones(tipoControl, df_salida):
    """
    Descripción: Añade pares de elementos que se consideran duplicados para tenerlo en cuenta a la hora de realizar los controles.

    Parámetros:
        · tipoControl -> La clave del tipo de control correspondiente a la tabla "TipoControles" de DQA
        · df_salida -> DataFrame al que se le añaden las excepciones.

    Referencias:
        - Control_DuplicadosProveedoresFAST.py
        - Control_DuplicadosServiciosFAST.py
        - Control_DuplicadosServiciosSF.py
    """

    # Se obtienen las excepciones del tipo 2 que hay en la tabla "Filtro" de DQA para el control correspondiente
    query_excep = """
            SELECT "Atributo1", "Valor1", "Atributo2", "Valor2"
            FROM "DQA"."Filtro" fil 
                JOIN "DQA"."Excepcion" exc 
                    ON fil."ClExcepcion" = exc."ClExcepcion"
                JOIN "DQA"."RelExcepcionControl" rel
                    ON rel."ClExcepcion" = exc."ClExcepcion"
                JOIN "DQA"."TiposControles" tip
                    ON tip."ClTipControl" = rel."ClTipControl"
                JOIN "DQA"."TiposExcepciones" tipex
                    ON tipex."ClTipoExcepcion" = exc."ClTipoExcepcion"
            WHERE tip."ClTipControl" = """ + str(tipoControl) + """
                AND tipex."ClTipoExcepcion" = 2
                """
    df_excep = query_DQA(query_excep)
    df_return = df_salida
    if df_excep.count > 0:
        list_excep = df_excep.collect()
        list_valor1 = []
        list_valor2 = []
        list_attr1 = []
        list_attr2 = []
        for i in range(len(list_excep)):
            list_valor1.append(list_excep[i]['Valor1'])
            list_valor2.append(list_excep[i]['Valor2'])
            list_attr1.append(list_excep[i]['Atributo1'])
            list_attr2.append(list_excep[i]['Atributo2'])

        df_salida.createOrReplaceTempView("table")
        
        for j in range(len(list_attr1)):
            if (list_attr2[j] == None):
                # Se tiene en cuenta a la hora de considerarlo duplicado
                df_return = df_salida.withColumn("isDuplicate", \
                    F.when(F.col(list_attr1[j]).isin(list_valor1[j]), 1).otherwise(df_salida.isDuplicate))
            else:
                # Se tiene en cuenta a la hora de considerarlo duplicado
                df_return = df_salida.withColumn("isDuplicate", \
                    F.when(( (F.col(list_attr1[j]).isin(list_valor1[j])) & (F.col(list_attr2[j]).isin(list_valor2[j])) ), 1).otherwise(df_salida.isDuplicate))
    return df_return

def fechas():
    """ 
    Descripción: Devuelve la fecha de hoy y la de ayer. En PRE tiene en cuenta que durante los fines de semana no se realiza la ejecución de los controles.
    
    Referencias:
        - Warning_VariabilidadDiariaATR.py
        - warning_sobrepaso_umbral_atributo.py
        - warning_wariabilidad_diaria.py
    
    """
    if config.ENTORNO == 'PRE':
        date_today = datetime.now()
        str_date = date_today.strftime('%Y-%m-%d')
        date_yesterday = datetime.strptime(str_date,'%Y-%m-%d') - timedelta(days=1)
        weekday_yesterday = date_yesterday.strftime('%A')
        if weekday_yesterday == 'Sunday':
            previousdate = date_yesterday - timedelta(days=2)
        else:
            previousdate = date_yesterday
        str_previousdate = previousdate.strftime('%Y-%m-%d')
        return str_date,str_previousdate
    if config.ENTORNO == 'PRO':
        date_today = datetime.now()
        date = date_today.strftime('%Y-%m-%d') 
        date_yesterday = datetime.strptime(date,'%Y-%m-%d') - timedelta(days=1)
        previousdate = date_yesterday.strftime('%Y-%m-%d')
        return date,previousdate

def write_ejecucionControl_OKNOK(OKNOK,NombreControl):
    """
    Descripción: Escribe en la tabla "Historico_EjecucionControlOKNOK" de DQA si el control se ha ejecutado correctamente o, por el contrario, algo ha fallado.

    Parámetros:
        · OKNOK -> OK si el control ha funcionado correctamente. NOK si el control ha fallado en algo.
        · NombreControl -> Nombre del control.

    Referencias:
        - Warning_Correct.py
        - Warning_SobrepasoUmbral.py
        - Warning_VariabilidadDiariaATR.py
        - warning_equipoPreventa.py
        - warning_mnc.py
        - warning_nulos_mandatories.py
        - warning_sobrepaso_umbral_atributo.py
        - warning_wariabilidad_diaria.py
    """
    df = spark.createDataFrame([[OKNOK,NombreControl]],schema=StructType([StructField("OKNOK", StringType()),StructField("ControlName", StringType())]))
    write_one_to_many_in_db(df,'"DQA"."Historico_EjecucionControlOKNOK"',39)


            
